import os
import re
from typing import Optional, List, Dict, Any

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None


class DocumentParser:
    """文档解析器 - 支持PDF/Word/Excel/TXT格式，含表格解析能力"""
    
    def parse(self, file_path: str, filename: str) -> str:
        """根据文件类型解析文档"""
        ext = filename.split('.')[-1].lower()
        
        if ext == 'pdf':
            return self._parse_pdf(file_path)
        elif ext in ['docx', 'doc']:
            return self._parse_docx(file_path)
        elif ext in ['xlsx', 'xls']:
            return self._parse_excel(file_path)
        elif ext == 'txt':
            return self._parse_txt(file_path)
        else:
            return self._parse_txt(file_path)
    
    def parse_with_tables(self, file_path: str, filename: str) -> Dict[str, Any]:
        """
        解析文档并提取表格数据
        返回: {text: str, tables: List[Dict], metadata: Dict}
        """
        ext = filename.split('.')[-1].lower()
        
        result = {
            'text': '',
            'tables': [],
            'metadata': {
                'filename': filename,
                'file_type': ext,
                'table_count': 0
            }
        }
        
        if ext == 'pdf':
            result['text'], result['tables'] = self._parse_pdf_with_tables(file_path)
        elif ext in ['xlsx', 'xls']:
            result['text'], result['tables'] = self._parse_excel_with_tables(file_path)
        elif ext in ['docx', 'doc']:
            result['text'] = self._parse_docx(file_path)
            result['tables'] = self._extract_tables_from_text(result['text'])
        elif ext == 'txt':
            result['text'] = self._parse_txt(file_path)
            result['tables'] = self._extract_tables_from_text(result['text'])
        elif ext in ['md', 'markdown']:
            result['text'] = self._parse_txt(file_path)
            result['tables'] = self._extract_markdown_tables(result['text'])
        
        result['metadata']['table_count'] = len(result['tables'])
        return result
    
    def extract_params(self, file_path: str, filename: str) -> List[Dict[str, Any]]:
        """
        从工艺参数文档中提取结构化参数
        返回: List[{param_name, value, unit, range, equipment, process}]
        """
        parsed = self.parse_with_tables(file_path, filename)
        params = []
        
        # 工艺参数关键词
        param_patterns = [
            (r'(温度|Temperature)[：:]*\s*([0-9.]+)\s*(°C|K|℃)', '温度', '°C'),
            (r'(压力|Pressure)[：:]*\s*([0-9.]+)\s*(mTorr|Torr|psi|pa|Pa)', '压力', None),
            (r'(流量|Flow)[：:]*\s*([0-9.]+)\s*(sccm|mL/min|L/min)', '流量', None),
            (r'(时间|Time|Duration)[：:]*\s*([0-9.]+)\s*(s|sec|min|分钟|秒)', '时间', None),
            (r'(功率|Power)[：:]*\s*([0-9.]+)\s*(W|kW|mW)', '功率', 'W'),
            (r'(转速|Speed|RPM)[：:]*\s*([0-9.]+)\s*(rpm|RPM)', '转速', 'rpm'),
            (r'(厚度|Thickness)[：:]*\s*([0-9.]+)\s*(nm|μm|mm|m)', '厚度', None),
            (r'(浓度|Concentration)[：:]*\s*([0-9.]+)\s*(%|ppm|mol/L)', '浓度', None),
        ]
        
        # 从表格中提取参数（表格结构：参数名 | 值 | 单位 | 标准范围）
        # 识别工艺参数类型的关键词
        param_keywords = ['温度', '压力', '流量', '时间', '功率', '转速', '厚度', '浓度', 
                          'Temperature', 'Pressure', 'Flow', 'Time', 'Power', 'Speed', 'Thickness',
                          '电压', '电流', '真空', '速度', '剂量', '能量', '温度']
        
        unit_keywords = ['°C', '℃', 'K', 'mTorr', 'Torr', 'psi', 'Pa', 'sccm', 'mL/min', 'L/min',
                         'W', 'kW', 'mW', 'rpm', 'RPM', 'nm', 'μm', 'mm', 'm', '%', 'ppm',
                         'mol/L', 's', 'sec', 'min', 'ms', 'mJ/cm²', 'V', 'A', 'Hz', 'g/L']
        
        for table in parsed.get('tables', []):
            rows = table.get('rows', [])
            for row in rows:
                # 跳过表头和分隔行
                row_text = '|'.join([str(c) for c in row])
                if any(skip in row_text for skip in ['---', '参数名称', '参数名', '名称', 'Parameter']):
                    continue
                
                # 在行中查找参数
                for cell in row:
                    cell_text = str(cell).strip()
                    if not cell_text:
                        continue
                    
                    # 检查是否包含参数关键词
                    is_param_name = any(kw in cell_text for kw in param_keywords)
                    if is_param_name:
                        # 找到参数名，尝试获取值
                        param_name = cell_text
                        
                        # 查找同行中的数值
                        for other_cell in row:
                            other_text = str(other_cell).strip()
                            # 检查是否是数值
                            if other_text and other_text != param_name:
                                # 数值匹配
                                num_match = re.match(r'^([0-9.]+)\s*(.*)$', other_text)
                                if num_match:
                                    value = num_match.group(1)
                                    unit = num_match.group(2).strip() if num_match.group(2) else ''
                                    
                                    # 确定参数类型
                                    param_type = '其他'
                                    for kw in param_keywords:
                                        if kw in param_name:
                                            if '温度' in param_name or 'Temperature' in param_name:
                                                param_type = '温度'
                                            elif '压力' in param_name or 'Pressure' in param_name:
                                                param_type = '压力'
                                            elif '流量' in param_name or 'Flow' in param_name:
                                                param_type = '流量'
                                            elif '时间' in param_name or 'Time' in param_name:
                                                param_type = '时间'
                                            elif '功率' in param_name or 'Power' in param_name:
                                                param_type = '功率'
                                            elif '转速' in param_name or 'Speed' in param_name or 'RPM' in param_name:
                                                param_type = '转速'
                                            elif '厚度' in param_name or 'Thickness' in param_name:
                                                param_type = '厚度'
                                            elif '浓度' in param_name or 'Concentration' in param_name:
                                                param_type = '浓度'
                                            else:
                                                param_type = '其他'
                                            break
                                    
                                    params.append({
                                        'param_name': param_name,
                                        'value': value,
                                        'unit': unit if unit else '',
                                        'param_type': param_type,
                                        'source': 'table_extraction',
                                        'context': row_text
                                    })
                                    break
        
        # 从文本中提取参数
        text = parsed['text']
        for pattern, param_type, default_unit in param_patterns:
            matches = re.finditer(pattern, text, re.IGNORECASE)
            for match in matches:
                param_name = match.group(1)
                value = match.group(2)
                unit = match.group(3) if len(match.groups()) > 2 else default_unit
                
                params.append({
                    'param_name': param_name,
                    'value': value,
                    'unit': unit or '',
                    'param_type': param_type,
                    'source': 'text_extraction',
                    'context': text[max(0, match.start()-20):match.end()+20]
                })
        
        # 从表格中提取参数
        for table in parsed['tables']:
            for row in table.get('rows', []):
                for cell in row:
                    cell_text = str(cell).strip()
                    for pattern, param_type, default_unit in param_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            match = re.search(pattern, cell_text, re.IGNORECASE)
                            if match:
                                params.append({
                                    'param_name': match.group(1),
                                    'value': match.group(2),
                                    'unit': match.group(3) if len(match.groups()) > 2 else default_unit,
                                    'param_type': param_type,
                                    'source': 'table_extraction',
                                    'context': cell_text
                                })
        
        # 去重
        seen = set()
        unique_params = []
        for p in params:
            key = (p['param_name'], p['value'], p['param_type'])
            if key not in seen:
                seen.add(key)
                unique_params.append(p)
        
        return unique_params
    
    def _parse_pdf(self, file_path: str) -> str:
        """解析PDF文件（基础文本提取）"""
        if PyPDF2 is None:
            return "[PDF解析不可用: PyPDF2未安装]"
        
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            text = f"[PDF解析错误: {str(e)}]"
        return text.strip()
    
    def _parse_pdf_with_tables(self, file_path: str):
        """解析PDF文件并提取表格"""
        text = ""
        tables = []
        
        if pdfplumber is None:
            # fallback to PyPDF2
            text = self._parse_pdf(file_path)
            tables = self._extract_tables_from_text(text)
            return text, tables
        
        try:
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    # 提取文本
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                    
                    # 提取表格
                    page_tables = page.extract_tables()
                    for table in page_tables:
                        if table and len(table) > 0:
                            tables.append({
                                'page': page_num + 1,
                                'rows': table,
                                'row_count': len(table)
                            })
        except Exception as e:
            text += f"\n[PDF表格解析错误: {str(e)}]"
        
        return text.strip(), tables
    
    def _parse_excel(self, file_path: str) -> str:
        """解析Excel文件"""
        if openpyxl is None:
            return "[Excel解析不可用: openpyxl未安装]"
        
        text = ""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"\n## Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(v) if v is not None else '' for v in row]
                    if any(v.strip() for v in row_values):
                        text += " | ".join(row_values) + "\n"
        except Exception as e:
            text = f"[Excel解析错误: {str(e)}]"
        return text.strip()
    
    def _parse_excel_with_tables(self, file_path: str):
        """解析Excel文件并提取表格"""
        text = ""
        tables = []
        
        if openpyxl is None:
            return "[Excel解析不可用]", []
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = f"\n## Sheet: {sheet_name}\n"
                
                current_table = []
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(v) if v is not None else '' for v in row]
                    if any(v.strip() for v in row_values):
                        sheet_text += " | ".join(row_values) + "\n"
                        current_table.append(row_values)
                
                text += sheet_text
                
                if current_table:
                    tables.append({
                        'sheet': sheet_name,
                        'rows': current_table,
                        'row_count': len(current_table)
                    })
        except Exception as e:
            text += f"\n[Excel表格解析错误: {str(e)}]"
        
        return text.strip(), tables
    
    def _parse_docx(self, file_path: str) -> str:
        """解析Word文件"""
        if DocxDocument is None:
            return "[Word解析不可用: python-docx未安装]"
        
        text = ""
        try:
            doc = DocxDocument(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
        except Exception as e:
            text = f"[Word解析错误: {str(e)}]"
        return text.strip()
    
    def _parse_txt(self, file_path: str) -> str:
        """解析文本文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='gbk') as file:
                    return file.read()
            except:
                return "[文本解析错误: 无法解码]"
        except Exception as e:
            return f"[文本解析错误: {str(e)}]"
    
    def _extract_markdown_tables(self, text: str) -> List[Dict]:
        """
        从Markdown文本中识别并提取表格
        Markdown表格格式: | col1 | col2 | 或 |---|---|
        """
        tables = []
        lines = text.split('\n')
        current_table = []
        in_table = False

        for line in lines:
            stripped = line.strip()
            # 检测是否是表格行：| cell1 | cell2 | ...
            if stripped.startswith('|') and stripped.endswith('|'):
                cells = [c.strip() for c in stripped.split('|') if c.strip() != '']
                # 跳过分隔行 |---|---|---|
                if cells and not all(re.match(r'^[-:]+$', c) for c in cells):
                    current_table.append(cells)
                    in_table = True
            else:
                if in_table and len(current_table) > 1:
                    tables.append({
                        'rows': current_table,
                        'row_count': len(current_table),
                        'source': 'markdown'
                    })
                current_table = []
                in_table = False

        if in_table and len(current_table) > 1:
            tables.append({
                'rows': current_table,
                'row_count': len(current_table),
                'source': 'markdown'
            })

        return tables

    def _extract_tables_from_text(self, text: str) -> List[Dict]:
        """
        从文本中识别表格（管道符分隔的表格）
        """
        tables = []
        lines = text.split('\n')
        current_table = []
        
        for line in lines:
            # 管道符分隔的表格行
            if '|' in line:
                cells = [c.strip() for c in line.split('|')]
                if len(cells) >= 2:
                    current_table.append(cells)
            else:
                if current_table and len(current_table) > 1:
                    tables.append({
                        'rows': current_table,
                        'row_count': len(current_table)
                    })
                current_table = []
        
        if current_table and len(current_table) > 1:
            tables.append({
                'rows': current_table,
                'row_count': len(current_table)
            })
        
        return tables
